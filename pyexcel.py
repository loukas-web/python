from openpyxl import *

wb = load_workbook("Grades.xlsx")
ws = wb.active

#print(ws["A1"].value)
#ws["A1"].value = "Full Name"
#wb.create_sheet("grades3")
print(wb.worksheets)

wb.save("Grades.xlsx")

for i in range(100):
    if i+i+1+i+2+i+3+i+4 == 54:
        print(i)